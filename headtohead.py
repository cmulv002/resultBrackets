from lxml import html

import requests
import string

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

#----------------------------------------#
#-----------------URL List---------------#
#----------------------------------------#
#These URLS are used to gather the head to head match data for all the
#matches played between top 100 players in 2018. This program currently requires
#the URL from each of the brackets to be included. This program does not
#currently support waterfall brackets or round robin style pools

#summit 6 - pools must be manually added to the spreadsheet on completion
summit6Bracket = 'https://smash.gg/tournament/smash-summit-6/events/melee-singles/brackets/254732'

#evo
evoTop8 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329225'
evoTop48 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329222'
evoH400 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/670331'
evoH401 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/670332'
evoH402 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/670333'
evoH403 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/670334'
evoH404 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/670335'
evoH405 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/670336'
evoH406 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/670337'
evoH407 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/670338'
evoI400 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/663136'
evoI401 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/663134'
evoI402 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/663134'
evoI403 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/663133'
evoI404 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/663132'
evoI405 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/663131'
evoI406 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/663128'
evoI407 = 'https://smash.gg/tournament/evo-2018/events/evo-2018-1/brackets/329220/663130'



#ltc
ltcTop8 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/317876'
ltcTop64 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/317875'
ltcA1 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649206'
ltcA2 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649209'
ltcA3 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649214'
ltcB1 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649213'
ltcB2 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649218'
ltcC1 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649216'
ltcC2 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649217'
ltcF1 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649211'
ltcF2 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649220'
ltcF3 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649210'
ltcG1 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/521338'
ltcG2 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649215'
ltcG3 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649219'
ltcH1 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649207'
ltcH2 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649212'
ltcH3 = 'https://smash.gg/tournament/kumite-in-texas-low-tier-city-6/events/melee-singles/brackets/219805/649208'

#CEO 2018
ceoTop24 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276726'
ceoB1 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599783'
ceoB2 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599786'
ceoB3 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599787'
ceoB4 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599788'
ceoB5 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599789'
ceoB6 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599790'
ceoB7 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599791'
ceoB8 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599792'
ceoB9 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599793'
ceoB10 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599784'
ceoB11 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599785'
ceoC1 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599794'
ceoC2 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599797'
ceoC3 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599798'
ceoC4 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599799'
ceoC5 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599800'
ceoC6 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599801'
ceoC7 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599802'
ceoC8 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599803'
ceoC9 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599804'
ceoC10 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599795'
ceoC11 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599796'
ceoD1 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599805'
ceoD2 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599807'
ceoD3 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599808'
ceoD4 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599809'
ceoD5 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599810'
ceoD6 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599811'
ceoD7 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599812'
ceoD8 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599813'
ceoD9 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599814'
ceoD10 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276724/599806'
ceoG1 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276725/599815'
ceoG7 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276725/599816'
ceoH1 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276725/599817'
ceoH7 = 'https://smash.gg/tournament/ceo-2018-fighting-game-championships/events/super-smash-bros-melee/brackets/276725/599818'
#DreamHack Austin 2018
dhaustinTop8 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/267121'
dhaustinTop48 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206535'
dhaustinA1 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504144'
dhaustinA2 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504152'
dhaustinA3 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504143'
dhaustinA4 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504145'
dhaustinB1 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504150'
dhaustinB2 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504147'
dhaustinB3 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504149'
dhaustinB4 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504154'
dhaustinC1 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504151'
dhaustinC2 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504153'
dhaustinC3 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504155'
dhaustinC4 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504142'
dhaustinD1 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504146'
dhaustinD2 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504156'
dhaustinD3 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504157'
dhaustinD4 = 'https://smash.gg/tournament/dreamhack-austin-2018/events/super-smash-bros-melee/brackets/206534/504148'

#Smash n’ Splash
snsTop64 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/263225'
snsG1 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583820'
snsG2 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583808'
snsG3 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583810'
snsG4 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583818'
snsH1 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583822'
snsH2 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583814'
snsH3 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583809'
snsH4 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583821'
snsI1 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583819'
snsI2 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583815'
snsI3 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583823'
snsI4 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583816'
snsJ1 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583811'
snsJ2 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583812'
snsJ3 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583813'
snsJ4 = 'https://smash.gg/tournament/smash-n-splash-4/events/melee-singles/brackets/265452/583817'



#Momocon 2018
momoI1 = 'https://smash.gg/tournament/momocon-2018-1/events/melee-singles/brackets/261806/578323'
momoJ1 = 'https://smash.gg/tournament/momocon-2018-1/events/melee-singles/brackets/261806/578324'
momoSunday = 'https://smash.gg/tournament/momocon-2018-1/events/melee-singles/brackets/261807'




#Goml 2018
gomlTop24 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189168'
gomlEE3 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189169/477080'
gomlEE4 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189169/477081'
gomlB1 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477047'
gomlB2 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477053'
gomlB3 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477058'
gomlB4 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477055'
gomlC1 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477052'
gomlC2 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477057'
gomlC3 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477062'
gomlC4 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477056'
gomlD1 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477048'
gomlD2 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477049'
gomlD3 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477061'
gomlD4 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477054'
gomlE1 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477060'
gomlE2 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477050'
gomlE3 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477051'
gomlE4 = 'https://smash.gg/tournament/get-on-my-level-2018-canadian-smash-championships/events/super-smash-bros-melee-singles/brackets/189167/477059'

#noodsnoodsnoodsoakland
noodsOakTop32 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/200172'
noodsOakA1 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/187824/493724'
noodsOakA2 = 'https://smash.gg/tournament/noods-noods-noods-oakland-edition-1/events/melee-singles/brackets/210329/509257'
noodsOakA3 = 'https://smash.gg/tournament/noods-noods-noods-oakland-edition-1/events/melee-singles/brackets/210329/509259'
noodsOakA4 = 'https://smash.gg/tournament/noods-noods-noods-oakland-edition-1/events/melee-singles/brackets/210329/509261'
noodsOakB1 = 'https://smash.gg/tournament/noods-noods-noods-oakland-edition-1/events/melee-singles/brackets/210329/509256'
noodsOakB2 = 'https://smash.gg/tournament/noods-noods-noods-oakland-edition-1/events/melee-singles/brackets/210329/509258'
noodsOakB3 = 'https://smash.gg/tournament/noods-noods-noods-oakland-edition-1/events/melee-singles/brackets/210329/509260'
noodsOakB4 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/187824/493722'

#flatiron3
flatiron3Top48 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/250880'
flatiron3A1 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/453121'
flatiron3A2 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559283'
flatiron3A3 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559285'
flatiron3A4 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559287'
flatiron3A5 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559292'
flatiron3A6 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559295'
flatiron3A7 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559291'
flatiron3A8 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559288'
flatiron3B1 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559281'
flatiron3B2 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559282'
flatiron3B3 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559284'
flatiron3B4 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559286'
flatiron3B5 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559289'
flatiron3B6 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559290'
flatiron3B7 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559293'
flatiron3B8 = 'https://smash.gg/tournament/flatiron-3/events/melee-singles/brackets/172876/559294'

#fullbloom4
fullbloom4Top48 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/205181'
fullbloom4A1 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502200'
fullbloom4A2 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502201'
fullbloom4A3 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502202'
fullbloom4A4 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502206'
fullbloom4B1 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502199'
fullbloom4B2 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502198'
fullbloom4B3 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502210'
fullbloom4B4 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502205'
fullbloom4C1 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502209'
fullbloom4C2 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502212'
fullbloom4C3 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502204'
fullbloom4C4 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502208'
fullbloom4D1 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/431113'
fullbloom4D2 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502211'
fullbloom4D3 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502203'
fullbloom4D4 = 'https://smash.gg/tournament/full-bloom-4/events/melee-singles/brackets/159450/502207'

#elgx
elgxTop8 = 'https://smash.gg/tournament/eglx-2018-smash/events/melee-singles-at-eglx-2018/brackets/220329'
elgxTop32 = 'https://smash.gg/tournament/eglx-2018-smash/events/melee-singles-at-eglx-2018/brackets/217930'
elgxA1 = 'https://smash.gg/tournament/eglx-2018-smash/events/melee-singles-at-eglx-2018/brackets/185941/472138'
elgxA2 = 'https://smash.gg/tournament/eglx-2018-smash/events/melee-singles-at-eglx-2018/brackets/185941/518764'
elgxA3 = 'https://smash.gg/tournament/eglx-2018-smash/events/melee-singles-at-eglx-2018/brackets/185941/518765'
elgxA4 = 'https://smash.gg/tournament/eglx-2018-smash/events/melee-singles-at-eglx-2018/brackets/185941/518766'
elgxB1 = 'https://smash.gg/tournament/eglx-2018-smash/events/melee-singles-at-eglx-2018/brackets/185941/518767'
elgxB2 = 'https://smash.gg/tournament/eglx-2018-smash/events/melee-singles-at-eglx-2018/brackets/185941/518768'
elgxB3 = 'https://smash.gg/tournament/eglx-2018-smash/events/melee-singles-at-eglx-2018/brackets/185941/518769'
elgxB4 = 'https://smash.gg/tournament/eglx-2018-smash/events/melee-singles-at-eglx-2018/brackets/185941/518770'


#smashvalley7
valley7Top48 = 'https://smash.gg/tournament/smash-valley-lucky-7/events/melee-singles/brackets/184298'
valley7A1 = 'https://smash.gg/tournament/smash-valley-lucky-7/events/melee-singles/brackets/178264/461185'
valley7A2 = 'https://smash.gg/tournament/smash-valley-lucky-7/events/melee-singles/brackets/178264/476296'
valley7A3 = 'https://smash.gg/tournament/smash-valley-lucky-7/events/melee-singles/brackets/178264/476297'
valley7A4 = 'https://smash.gg/tournament/smash-valley-lucky-7/events/melee-singles/brackets/178264/476298'
valley7B1 = 'https://smash.gg/tournament/smash-valley-lucky-7/events/melee-singles/brackets/178264/476299'
valley7B2 = 'https://smash.gg/tournament/smash-valley-lucky-7/events/melee-singles/brackets/178264/476300'
valley7B3 = 'https://smash.gg/tournament/smash-valley-lucky-7/events/melee-singles/brackets/178264/476301'
valley7B4 = 'https://smash.gg/tournament/smash-valley-lucky-7/events/melee-singles/brackets/178264/476302'


#noodsnoodsnoodsmelee
noodsmeleeTop32 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/200172'
noodsmeleeA1 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/187824/493724'
noodsmeleeA2 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/187824/493723'
noodsmeleeA3 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/187824/492544'
noodsmeleeA4 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/187824/492543'
noodsmeleeB1 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/187824/474764'
noodsmeleeB2 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/187824/493760'
noodsmeleeB3 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/187824/492542'
noodsmeleeB4 = 'https://smash.gg/tournament/noods-noods-noods-melee-edition/events/melee-singles/brackets/187824/493722'

#genesis5
genesisTop8 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190488'
genesisTop64 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190487'
genesisH13 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478930'
genesisH14 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478931'
genesisH15 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478932'
genesisH16 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478934'
genesisI1 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478942'
genesisI2 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478938'
genesisI3 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478940'
genesisI4 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478927'
genesisI5 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478937'
genesisI6 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478944'
genesisJ1 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478933'
genesisJ2 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478936'
genesisJ3 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478943'
genesisJ4 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478935'
genesisJ5 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478939'
genesisJ6 = 'https://smash.gg/tournament/genesis-5/events/melee-singles/brackets/190486/478941'

#optic arena 
opticBracket = 'https://smash.gg/tournament/optic-arena/events/super-smash-bros-melee-singles/brackets/270792'
opticA1 = 'https://smash.gg/tournament/optic-arena/events/super-smash-bros-melee-singles/brackets/199992/591076'
opticA2 = 'https://smash.gg/tournament/optic-arena/events/super-smash-bros-melee-singles/brackets/199992/591080'
opticB1 = 'https://smash.gg/tournament/optic-arena/events/super-smash-bros-melee-singles/brackets/199992/591079'
opticB2 = 'https://smash.gg/tournament/optic-arena/events/super-smash-bros-melee-singles/brackets/199992/591075'
opticC1 = 'https://smash.gg/tournament/optic-arena/events/super-smash-bros-melee-singles/brackets/199992/591078'
opticC2 = 'https://smash.gg/tournament/optic-arena/events/super-smash-bros-melee-singles/brackets/199992/591074'
opticD1 = 'https://smash.gg/tournament/optic-arena/events/super-smash-bros-melee-singles/brackets/199992/493038'
opticD2 = 'https://smash.gg/tournament/optic-arena/events/super-smash-bros-melee-singles/brackets/199992/591077'





#bigger balc
balcBracket = 'https://smash.gg/tournament/the-even-bigger-balc/events/melee-singles/brackets/283305'
balcA1 = 'https://smash.gg/tournament/the-even-bigger-balc/events/melee-singles/brackets/204239/608490'
balcA2 = 'https://smash.gg/tournament/the-even-bigger-balc/events/melee-singles/brackets/204239/608492'
balcB1 = 'https://smash.gg/tournament/the-even-bigger-balc/eevents/melee-singles/brackets/204239/608488'
balcB2 = 'https://smash.gg/tournament/the-even-bigger-balc/events/melee-singles/brackets/204239/608489'
balcC1 = 'https://smash.gg/tournament/the-even-bigger-balc/events/melee-singles/brackets/204239/608486'
balcC2 = 'https://smash.gg/tournament/the-even-bigger-balc/events/melee-singles/brackets/204239/608491'
balcD1 = 'https://smash.gg/tournament/the-even-bigger-balc/events/melee-singles/brackets/204239/608487'
balcD2 = 'https://smash.gg/tournament/the-even-bigger-balc/events/melee-singles/brackets/204239/500984'

#dream hack summer 2018 in sweden
dhswedenBracket = 'https://smash.gg/tournament/dreamhack-summer-2018/events/dreamhack-smash-championship-melee-singles/brackets/206821'

#smash factor 7 
sf7Top24 = 'https://smash.gg/tournament/smash-factor-7/events/melee-singles/brackets/307559'
sf7PoolI1 = 'https://smash.gg/tournament/smash-factor-7/events/melee-singles/brackets/307557/637399'
sf7PoolI2 = 'https://smash.gg/tournament/smash-factor-7/events/melee-singles/brackets/307557/637383'

#Overlords of Orlando3
overlords3Bracket = 'https://smash.gg/tournament/overlords-of-orlando-3-three-games-one-roof/events/melee-singles/brackets/269586'

#Aurora Blitz
auroraBracket = 'https://smash.gg/tournament/aurora-blitz-a-super-smash-bros-melee-and-wii-u/events/melee-singles/brackets/317787'



#OmegaII 
omega2Bracket = 'https://smash.gg/tournament/omega-ii/events/melee-singles/brackets/202824'
#Runback2018
runback2018Top16 = 'https://smash.gg/tournament/runback-2018/events/melee-singles/brackets/303969'
runback2018Top48 = 'https://smash.gg/tournament/runback-2018/events/melee-singles/brackets/296932'
#talking stick arizona
talkingstickTop8 = 'https://smash.gg/tournament/esports-arizona-talking-stick-resort-4/events/super-smash-bros-melee-singles/brackets/277369'
talkingstickTop32 = 'https://smash.gg/tournament/esports-arizona-talking-stick-resort-4/events/super-smash-bros-melee-singles/brackets/277366'
#poi poundaz
poipoundazTop16 = 'https://smash.gg/tournament/poi-poundaz/events/melee-singles/brackets/182821'
#combo breaker 2018
cb2018 = 'https://smash.gg/tournament/combo-breaker-2018-1/events/super-smash-bros-melee/brackets/262744/580080'




#valhalla
valhallaTop64 = 'https://smash.gg/tournament/valhalla/events/melee-singles/brackets/188843'

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
#sheet['A1'] = 200
#sheetPosition = 'A1'


    
    

def tourneyFunTop8(url): #FOR TOP 8 ONLY
    page = requests.get(url) #sets page equal to the url
    #print(url) 
    root = html.fromstring(page.content) #sets root equal to the page content 
    loserList = []
    winnerList = []
    playerList = []
    name1 = []
    name2 = []
    name3 = []
    name4 = []
    countList = []
    incrementNameList = []
    winnerCheck = None

    
    falseStrTxt = False
    falseStrTxt2 = False
    falseStrTxt3 = False
    noReport = False

    x = 0
    z = 0
    
    
    
    for el in root.cssselect('div.match-affix-wrapper'): #div.match-affix-wrapper is the match selector
        

        for elemCheck1 in el.cssselect('div.matchSectionWrapper'): 
            incrementNameList.append(elemCheck1.text_content())
            
        for elemCheck3 in el.cssselect('div.match-player.winner'): #
            for elemCheck4 in elemCheck3.cssselect('i.fa.fa-check.text-success'): #checks if there is a non-reported match
                winnerCheck = elemCheck3.text_content()
                #print(winnerCheck)
                noReport = True
            
        
        #print(incrementNameList)
        
        if(incrementNameList[0] == winnerCheck and noReport == True): #if score is not reported
            winnerList.append(winnerCheck+str(0))
            loserList.append(incrementNameList[1]+str(0))
            
            noReport = True
            falseStrTxt = True
            falseStrTxt2 = True
            
        elif(incrementNameList[0] != winnerCheck and noReport == True): #if score is reported
            loserList.append(incrementNameList[0]+str(0))
            winnerList.append(incrementNameList[1]+str(0))
            
            noReport = True
            falseStrTxt = True
            falseStrTxt2 = True
            
        incrementNameList = []
                    
                        
        
        for elem1 in el.cssselect('div.match-player.winner'): #
            
            if falseStrTxt == False:
                winnerList.append(elem1.text_content())
            falseStrTxt = False
            
            
                    
                
             
        
            
        for elem2 in el.cssselect('div.match-player.loser'):
            if falseStrTxt2 == False:
                loserList.append(elem2.text_content())
            falseStrTxt2 = False
        falseStrTxt = False
        falseStrTxt2 = False
        noReport = False    
        incrementNameList = []
        winnercheck = None
        x = x+1


    

    x = 1
    y = 1

    
    

    for d in range(0, len(winnerList)):  #removes game score from string
        winnerList[d] = winnerList[d][:-1]
        
        
        loserList[d] = loserList[d][:-1]
        #for getting only game count,
        #winnerList[d] = winnerList[d][-1:]
        #loserList[d] = loserList[d][-1:]



    x = 1
    y = 1

   

  
    sheetWriteFun100(winnerList, loserList) #function for writing to the sheet
    


    

def sheetWriteFun100(winnerList, loserList):
    global calcSheetWinner
    global calcSheetLoser
    global top100List
    deleteFlag2 = False
    deleteFlag1 = False
    newWinnerList = []
    newLoserList = []
    
    
    
    aCoord = 'A'
    colLetter = 'A'
    
    k = 0
    for j in range(2, 102):
        colLetter = get_column_letter(j)
        
        top100Coord1  = aCoord + str(j)
        top100Coord2 = colLetter + '1'
        sheet[top100Coord1] = top100List[k]
        sheet[top100Coord2] = top100List[k]
        k = k + 1
    k = 0

    

    winnerList = [el.replace('\xa0',' ') for el in winnerList]
    loserList = [el.replace('\xa0',' ') for el in loserList]

    

    for d in range(0, len(winnerList)):
        winnerList[d] = winnerList[d][1:]
        loserList[d] = loserList[d][1:]

    

    for e in range(0, len(winnerList)):
        for f in range(0, len(top100List)):
            if loserList[e] == top100List[f]:
                deleteFlag2 = True
            if winnerList[e] == top100List[f]:
                deleteFlag1 = True
                
        if (deleteFlag1 == True) and (deleteFlag2 == True):
            newWinnerList.append(winnerList[e])
            newLoserList.append(loserList[e])
            
        #else:
            #print("Delete ", winnerList[e], " beat ", loserList[e])
           
        deleteFlag2 = False
        deleteFlag1 = False

    
    
    
    #print(winnerList)
    #print(loserList)

    calcSheetWinner = calcSheetWinner + newWinnerList
    calcSheetLoser = calcSheetLoser + newLoserList
    #for x in range(0, len(calcSheetWinner)):
        #print(str(x+1), calcSheetWinner[x], " beat ", calcSheetLoser[x])
    

    #print(calcSheetWinner)
    #print(calcSheetLoser)
    

    
    

def sheetSetFun(winnerList, loserList):
    global top100List
    global calcSheetLoser
    global calcSheetWinner
    #print(winnerList)
    #print(loserList)
    winnerListNew = []
    fixerFlag1 = False
    fixerFlag2 = False
    winnerCoordNum = '1'
    loserCoordNum = '1'
    winnerCoordinate1 = 'A1'
    loserCoordinate1 = 'A1'
    winnerCoordinate2 = 'A1'
    loserCoordinate2 = 'A1'
    scoreshow = ''
    splitList = []
    winnerFlag1 = False
    loserFlag1 = False
    in100Test = False
    #print(winnerList)
    #print(loserList)
    winnerLength = len(winnerList)
    loserLength = len(loserList)
    sheetLength = winnerLength+loserLength
    endSheet = "A"+str(sheetLength)

    
    
    s = 0
    t = 0
    deleteFlag1 = False
    deleteFlag2 = False
    deleteFlag3 = False
    deleteList1 = []
    deleteList2 = []


    

    
    
    #print(winnerList)
    #print(loserList)

    #print(winnerLength, '-', loserLength)
    newWinnerList = []
    newLoserList = []
    
    

   
    
    #print(newWinnerList)
    #print(newLoserList)
    

    #for x in range(0, len(newWinnerList)):
        #print(x, newWinnerList[x], "beat", newLoserList[x])
        
                
    
    #for x in range(0, len(newWinnerList)):
        #print(str(x+1), winnerList[x], " beat ", loserList[x])
            
    winnerLength = len(winnerList)
    #print(winnerLength)
    loserLength = len(loserList)
    #print(loserLength)

    #print(loserList)

    scoreShow = None
        
    for a in range(0, winnerLength):
        for colCellObj in sheet['A2':'A101']:
            for cellObj1 in colCellObj:
                #print(str(cellObj1.value), '-', str(winnerList[a]))
                if cellObj1.value ==  winnerList[a]:
                    winnerCoordinate = cellObj1.coordinate
                    winnerCoordNum = winnerCoordinate.replace("A","")
                    #print(cellObj1.coordinate, winnerList[a], " Winner ")
                    fixerFlag1 = True
                    s=s+1
                    
                if cellObj1.value == loserList[a]:
                    loserCoordinate = cellObj1.coordinate
                    loserCoordNum = loserCoordinate.replace("A","")
                    #print(cellObj1.coordinate, loserList[a], " Loser ")
                    fixerFlag2 = True
                    t = t+1
                    
                    
                    
                #------SCOREDISP-------#
                #print(a)
        
            
            
        if fixerFlag1 == True and fixerFlag2 == True:
                winnerY = get_column_letter(int(winnerCoordNum))
                #print(winnerY, winnerCoordNum)
                loserY = get_column_letter(int(loserCoordNum))
                #print(loserY, loserCoordNum)
                fourCoordTwo = winnerY + loserCoordNum
                
                
                fourCoordOne = loserY + winnerCoordNum
                #print(fourCoordOne, fourCoordTwo)
                #print(fourCoordTwo)
                #print(fourCoordOne)
                fixerFlag1 == False
                fixerFlag2 == False
                
                #print(fourCoordOne, fourCoordTwo)
                #FOURCOORD1 is WINNER +
                #FOURCOORD2 is LOSER +
                scoreShow = sheet[fourCoordOne].value
                
                #print(scoreShow)
                if scoreShow is None:
                    
                    sheet[fourCoordOne] = '1-0'
                    sheet[fourCoordTwo] = '0-1'
                    #print(sheet[fourCoordTwo].value)
                else:
                    #print(scoreShow)
                    r = str(scoreShow)
                    splitList = r.split("-")
                    #print(splitList)
                    
                    ogWinScore = splitList[0]
                    ogWinScore = int(ogWinScore)
                    ogLoseScore = splitList[1]
                    ogWinScore = str(ogWinScore+1)
                    sheet[fourCoordOne] = ogWinScore + '-' + ogLoseScore
                    sheet[fourCoordTwo] = ogLoseScore + '-' + ogWinScore
        elif fixerFlag1 == True and fixerFlag2 == False:
            
            fixerFlag1 == False
            fixerFlag2 == False
        elif fixerFlag1 == False and fixerFlag2 == True:
            fixerFlag1 == False
            fixerFlag2 == False
            

                    #----------------------#
                   
    #for c in range(0, len(winnerList)-1):
        #print(calcSheetWinner[c], " beat ", calcSheetLoser[c])
    
        
    



#--------Main---------#
#---------------------#
            
calcSheetWinner = [] #this variable stores all of the winners across the brackets
calcSheetLoser = [] #this variable stores all of the losers across the bracket
tournamentVar = 'currently empty' #this stores the name of the current tournament 

#Top 100 List is a list of the SSBMRank 2017 Top 100 in order from 1-100
top100List = ['Liquid Hungrybox','[A] Armada', 'C9 Mang0', 'FOX MVG Mew2King','PG Plup',
              'TSM Leffen','Tempo Axe','Wizzrobe','CLG. SFAT', 'Tempo S2J', 'TL Chudat',
              'Balance Druggedfox', 'IMT Shroomed','Duck','Crush','OG Swedish Delight',
              'DIGNITAS Lucky', 'G2 Westballz', 'MSF La Luna', 'ALG n0ne', 'CLG. PewPewU',
              'PG Zain', 'DIGNITAS HugS', 'VGBC aMSa', 'beast coast MikeHaze', 'lloD',
              'dizzkidboogie', 'RB Ice','KirbyKaze', 'SS Colbol', 'vL Trif', 'Free Agent Ryan Ford',
              'Balance Syrox', 'Balance KJH', 'Nintendude', 'Rishi', 'Bladewise', 'Prince F. Abu',
              'Asterion Amsah', 'SD Santi', 'Professor Pro', 'FKA Slox', '20GX Gahtzu',
              '[A] Android', 'Hax', 'Mirage Kage', 'Captain Smuckers', 'MacD', 'Ginger',
              'Laudandus', 'GHQ Kels', 'Captain Faceroll', 'TNC ARMY', 'Junebug', 'Medz',
              'Abate', 'Liquid Chillindude', 'Squid', 'Azusa', 'Overtriforce', 'Spark',
              'Entropy Luigi Ka-Master', 'Trulliam', 'Cacutar', 'Redd', 'BOXR Zhu',
              'Darktama', 'Drephen', 'FatGoku', 'Kalamazhu', 'LFG Eddy Mexico',
              'Drunksloth', 'Cal', 'CLN AbsentPage', 'MVG | King Momo', 'DJ Nintendo',
              'homemadewaffles', 'lovage', 'OES MilkMan', 'Tai', 'Rik', 'TA Rocky',
              'Jerry', 'iBDW', 'RCS KPAN', 'FX Uncle mojo', 'FSBR $mike',
              'Sloth Darktooth', 'Zgetto', 'WLG Blea Gelo', 'Kaeon',  'Bobby Frizz',
              'Iceman', 'L', 'TA NMW', 'Moky', 'Legend', 'Zealous5000', 'Ralph', 'MnT 2saint']

#print(top100List) 


'''
#top 8 list for testing purposes
tourneyFunTop8(ltcTop8)
tourneyFunTop8(ltcTop64)
tourneyFunTop8(ceoTop24)
tourneyFunTop8(dhaustinTop8)
tourneyFunTop8(dhaustinTop48)
tourneyFunTop8(snsTop64)
tourneyFunTop8(momoSunday)
tourneyFunTop8(gomlTop24)
tourneyFunTop8(evoTop8)
tourneyFunTop8(evoTop48)
tourneyFunTop8(flatiron3Top48)
tourneyFunTop8(fullbloom4Top48)
tourneyFunTop8(elgxTop8)
tourneyFunTop8(elgxTop32)
tourneyFunTop8(genesisTop8)
tourneyFunTop8(genesisTop64)
tourneyFunTop8(valhallaTop64)
'''


#ltc6
tournamentVar = 'Low Tier City 6'
tourneyFunTop8(ltcTop8)
tourneyFunTop8(ltcTop64)
tourneyFunTop8(ltcA1)
tourneyFunTop8(ltcA2)
tourneyFunTop8(ltcA3)
tourneyFunTop8(ltcB1)
tourneyFunTop8(ltcB2)
tourneyFunTop8(ltcC1)
tourneyFunTop8(ltcC2)
tourneyFunTop8(ltcF1)
tourneyFunTop8(ltcF2)
tourneyFunTop8(ltcF3)
tourneyFunTop8(ltcG1)
tourneyFunTop8(ltcG2)
tourneyFunTop8(ltcG3)
tourneyFunTop8(ltcH1)
tourneyFunTop8(ltcH2)
tourneyFunTop8(ltcH3)


#ceo
tournamentVar = 'CEO 2018'
tourneyFunTop8(ceoTop24)
tourneyFunTop8(ceoB1)
tourneyFunTop8(ceoB2)
tourneyFunTop8(ceoB3)
tourneyFunTop8(ceoB4)
tourneyFunTop8(ceoB5)
tourneyFunTop8(ceoB6)
tourneyFunTop8(ceoB7)
tourneyFunTop8(ceoB8)
tourneyFunTop8(ceoB9)
tourneyFunTop8(ceoB10)
tourneyFunTop8(ceoB11)
tourneyFunTop8(ceoC1)
tourneyFunTop8(ceoC2)
tourneyFunTop8(ceoC3)
tourneyFunTop8(ceoC4)
tourneyFunTop8(ceoC5)
tourneyFunTop8(ceoC6)
tourneyFunTop8(ceoC7)
tourneyFunTop8(ceoC8)
tourneyFunTop8(ceoC9)
tourneyFunTop8(ceoC10)
tourneyFunTop8(ceoC11)
tourneyFunTop8(ceoD1)
tourneyFunTop8(ceoD2)
tourneyFunTop8(ceoD3)
tourneyFunTop8(ceoD4)
tourneyFunTop8(ceoD5)
tourneyFunTop8(ceoD6)
tourneyFunTop8(ceoD7)
tourneyFunTop8(ceoD8)
tourneyFunTop8(ceoD9)
tourneyFunTop8(ceoD10)
tourneyFunTop8(ceoG1)
tourneyFunTop8(ceoG7)
tourneyFunTop8(ceoH1)
tourneyFunTop8(ceoH7)

#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []


#dreamhackaustin
tournamentVar = 'Dream Hack Austin 2018'
tourneyFunTop8(dhaustinTop8)
tourneyFunTop8(dhaustinTop48)
tourneyFunTop8(dhaustinA1)
tourneyFunTop8(dhaustinA2)
tourneyFunTop8(dhaustinA3)
tourneyFunTop8(dhaustinA4)
tourneyFunTop8(dhaustinB1)
tourneyFunTop8(dhaustinB2)
tourneyFunTop8(dhaustinB3)
tourneyFunTop8(dhaustinB4)
tourneyFunTop8(dhaustinC1)
tourneyFunTop8(dhaustinC2)
tourneyFunTop8(dhaustinC3)
tourneyFunTop8(dhaustinC4)
tourneyFunTop8(dhaustinD1)
tourneyFunTop8(dhaustinD2)
tourneyFunTop8(dhaustinD3)
tourneyFunTop8(dhaustinD4)

#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []

#smash n' splash
tournamentVar = 'Smash and Splash 4'
tourneyFunTop8(snsTop64)
tourneyFunTop8(snsG1)
tourneyFunTop8(snsG2)
tourneyFunTop8(snsG3)
tourneyFunTop8(snsG4)
tourneyFunTop8(snsH1)
tourneyFunTop8(snsH2)
tourneyFunTop8(snsH3)
tourneyFunTop8(snsH4)

#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []


tourneyFunTop8(snsI1)
tourneyFunTop8(snsI2)
tourneyFunTop8(snsI3)
tourneyFunTop8(snsI4)
tourneyFunTop8(snsJ1)
tourneyFunTop8(snsJ2)
tourneyFunTop8(snsJ3)
tourneyFunTop8(snsJ4)

tournamentVar = 'Momocon'
tourneyFunTop8(momoI1)
tourneyFunTop8(momoJ1)
tourneyFunTop8(momoSunday)

#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []

#goml 2018
tournamentVar = 'Get on My Level 2018'
tourneyFunTop8(gomlTop24)
tourneyFunTop8(gomlEE3)
tourneyFunTop8(gomlEE4)
tourneyFunTop8(gomlB1)
tourneyFunTop8(gomlB2)
tourneyFunTop8(gomlB3)
tourneyFunTop8(gomlB4)
tourneyFunTop8(gomlC1)
tourneyFunTop8(gomlC2)
tourneyFunTop8(gomlC3)
tourneyFunTop8(gomlC4)
tourneyFunTop8(gomlD1)
tourneyFunTop8(gomlD2)
tourneyFunTop8(gomlD3)
tourneyFunTop8(gomlD4)
tourneyFunTop8(gomlE1)
tourneyFunTop8(gomlE2)
tourneyFunTop8(gomlE3)
tourneyFunTop8(gomlE4)

tournamentVar = 'Evo 2018'
tourneyFunTop8(evoTop8)
tourneyFunTop8(evoTop48)
tourneyFunTop8(evoH400)
tourneyFunTop8(evoH401)
tourneyFunTop8(evoH402)
tourneyFunTop8(evoH403)
tourneyFunTop8(evoH404)
tourneyFunTop8(evoH405)
tourneyFunTop8(evoH406)
tourneyFunTop8(evoH407)
tourneyFunTop8(evoI400)
tourneyFunTop8(evoI401)
tourneyFunTop8(evoI402)
tourneyFunTop8(evoI403)
tourneyFunTop8(evoI404)
tourneyFunTop8(evoI405)
tourneyFunTop8(evoI406)
tourneyFunTop8(evoI407)


#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []


tournamentVar = 'Flatiron 3'
tourneyFunTop8(flatiron3Top48)
tourneyFunTop8(flatiron3A1)
tourneyFunTop8(flatiron3A2)
tourneyFunTop8(flatiron3A3)
tourneyFunTop8(flatiron3A4)
tourneyFunTop8(flatiron3A5)
tourneyFunTop8(flatiron3A6)
tourneyFunTop8(flatiron3A7)
tourneyFunTop8(flatiron3A8)
tourneyFunTop8(flatiron3B1)
tourneyFunTop8(flatiron3B2)
tourneyFunTop8(flatiron3B3)
tourneyFunTop8(flatiron3B4)
tourneyFunTop8(flatiron3B5)
tourneyFunTop8(flatiron3B6)
tourneyFunTop8(flatiron3B7)
tourneyFunTop8(flatiron3B8)

#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []

tournamentVar = 'Full Bloom 4'
tourneyFunTop8(fullbloom4Top48)
tourneyFunTop8(fullbloom4A1)
tourneyFunTop8(fullbloom4A2)
tourneyFunTop8(fullbloom4A3)
tourneyFunTop8(fullbloom4A4)
tourneyFunTop8(fullbloom4B1)
tourneyFunTop8(fullbloom4B2)
tourneyFunTop8(fullbloom4B3)
tourneyFunTop8(fullbloom4B4)
tourneyFunTop8(fullbloom4C1)
tourneyFunTop8(fullbloom4C2)
tourneyFunTop8(fullbloom4C3)
tourneyFunTop8(fullbloom4C4)
tourneyFunTop8(fullbloom4D1)
tourneyFunTop8(fullbloom4D2)
tourneyFunTop8(fullbloom4D3)
tourneyFunTop8(fullbloom4D4)

#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []

tournamentVar = 'ELGX 2018'
tourneyFunTop8(elgxTop8)
tourneyFunTop8(elgxTop32)
tourneyFunTop8(elgxA1)
tourneyFunTop8(elgxA2)
tourneyFunTop8(elgxA3)
tourneyFunTop8(elgxA4)
tourneyFunTop8(elgxB1)
tourneyFunTop8(elgxB2)
tourneyFunTop8(elgxB3)
tourneyFunTop8(elgxB4)

#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []

tournamentVar = 'Smash Valley Lucky #7'
tourneyFunTop8(valley7Top48)
tourneyFunTop8(valley7A1)
tourneyFunTop8(valley7A2)
tourneyFunTop8(valley7A3)
tourneyFunTop8(valley7A4)
tourneyFunTop8(valley7B1)
tourneyFunTop8(valley7B2)
tourneyFunTop8(valley7B3)
tourneyFunTop8(valley7B4)

#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []


#Noods Noods Noods Melee Edition
tournamentVar = 'Noods Melee Edition'
tourneyFunTop8(noodsmeleeTop32)
tourneyFunTop8(noodsmeleeA1)
tourneyFunTop8(noodsmeleeA2)
tourneyFunTop8(noodsmeleeA3)
tourneyFunTop8(noodsmeleeA4)
tourneyFunTop8(noodsmeleeB1)
tourneyFunTop8(noodsmeleeB2)
tourneyFunTop8(noodsmeleeB3)
tourneyFunTop8(noodsmeleeB4)

#Noods noods noods Oakland
tournamentVar = 'Noods Oakland'
tourneyFunTop8(noodsOakTop32)
tourneyFunTop8(noodsOakA1)
tourneyFunTop8(noodsOakA2)
tourneyFunTop8(noodsOakA3)
tourneyFunTop8(noodsOakA4)
tourneyFunTop8(noodsOakB1)
tourneyFunTop8(noodsOakB2)
tourneyFunTop8(noodsOakB3)
tourneyFunTop8(noodsOakB4)

#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []


#Genesis 5
tournamentVar = 'Genesis 5'
tourneyFunTop8(genesisTop8)
tourneyFunTop8(genesisTop64)
tourneyFunTop8(genesisH13)
tourneyFunTop8(genesisH14)
tourneyFunTop8(genesisH15)
tourneyFunTop8(genesisH16)
tourneyFunTop8(genesisI1)
tourneyFunTop8(genesisI2)
tourneyFunTop8(genesisI3)
tourneyFunTop8(genesisI4)
tourneyFunTop8(genesisI5)
tourneyFunTop8(genesisI6)
tourneyFunTop8(genesisJ1)
tourneyFunTop8(genesisJ2)
tourneyFunTop8(genesisJ3)
tourneyFunTop8(genesisJ4)
tourneyFunTop8(genesisJ5)
tourneyFunTop8(genesisJ6)

#Valhalla
tournamentVar = 'Valhalla'
tourneyFunTop8(valhallaTop64)

#optic arena
tournamentVar = 'Optic Arena'
tourneyFunTop8(opticBracket)
tourneyFunTop8(opticA1)
tourneyFunTop8(opticA2)
tourneyFunTop8(opticB1)
tourneyFunTop8(opticB2)
tourneyFunTop8(opticC1)
tourneyFunTop8(opticC2)
tourneyFunTop8(opticD1)
tourneyFunTop8(opticD2)

#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []


#dream hack summer 2018 sweden
tournamentVar = 'Dreamhack Summer 2018 Sweden'
tourneyFunTop8(dhswedenBracket)

#smash factor 7
tournamentVar = 'Smash Factor 7'
tourneyFunTop8(sf7Top24)
tourneyFunTop8(sf7PoolI1)
tourneyFunTop8(sf7PoolI2)

#overlords
tournamentVar = 'Overlords of Orlando 3'
tourneyFunTop8(overlords3Bracket)

#aurora
tournamentVar = 'Aurora'
tourneyFunTop8(auroraBracket)

#sheetSetFun(calcSheetWinner, calcSheetLoser)
#calcSheetWinner = []
#calcSheetLoser = []


#lots of tourneys
tournamentVar = 'Omega II'
tourneyFunTop8(omega2Bracket)
tournamentVar = 'Runback 2018'
tourneyFunTop8(runback2018Top16)
tourneyFunTop8(runback2018Top48)
tournamentVar = 'Talking Stick'
tourneyFunTop8(talkingstickTop8)
tourneyFunTop8(talkingstickTop32)
tournamentVar = 'Poi Poundaz'
tourneyFunTop8(poipoundazTop16)
tournamentVar = 'Combo Breaker 2018'
tourneyFunTop8(cb2018)
tournamentVar = 'Summit 6'
tourneyFunTop8(summit6Bracket)


sheetSetFun(calcSheetWinner, calcSheetLoser)
calcSheetWinner = []
calcSheetLoser = []


#wb.save('writeBrackets.xlsx')
wb.save('writeBrackets.xlsx')


#---------------------#
#---------END---------#
#---------------------#



